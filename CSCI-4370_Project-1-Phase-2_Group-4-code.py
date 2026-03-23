# -*- coding: utf-8 -*-
"""
Data Mining Project 1, phase 2
Authors : Miles Glover, Madison Nicholson, Victory Orobosa

Programed in python

requires openpyxl library to function (only way to read xlsx files)
pip install openpyxl will suffice if openpyxl is missing from system

"""

from openpyxl import load_workbook

#Part 1 functions

#loads the exel file using openpyxl
def load_XLSX(file_name):

    print("Loading file")

    dataset = []

    wb = load_workbook(file_name)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    header = list(rows[0])

    #construct dataset row-by-row
    for row in rows[1:]:
        dataset.append(list(row))


    print("File loaded\n")

    return dataset,header

#function soley exists to remove the first attribute to each data row
def process_dataset(dataset):
    
    for row in dataset:
        row.pop(0)
    
    return dataset
    
#Part 2 functions
#function uses the test and training sets to perform knn with a desingated k. Jaccard distance is used for neighbor comparisons.
def perform_KNN(train,test, k): 

    print("Performing KNN")

    real_predicted = []
    
    for test_row in test:

        neighbors_distances = []

        #find k nearest neighbors
        for train_index, train_row in enumerate(train):

            #jaccard distance calculation
            p = 0
            q = 0
            r = 0

            #find the jaccard p,q,r values between the test and train rows
            for i in range(1, len(train_row)):
                Q = test_row[i]
                R = train_row[i]

                if Q == 1 and R == 1:   #same 1's so p increases
                    p += 1
                elif Q == 1 and R == 0: #test wins so q increases
                    q += 1
                elif Q == 0 and R == 1: #train wins so r increases
                    r += 1

            bottom = p + q + r

            #in case of no simularity, maximum distance achieved
            if bottom == 0:
                dist = 1.0
            
            #jaccard distance = 1 - jaccard simularity
            else:
                dist = 1-(p/bottom)

            #add distance to list
            neighbors_distances.append([train_index, dist])

        
        #sort neighbors_distances from lowest distance to highest
        #used classic selection sort
        for i in range(len(neighbors_distances)):
            min_index = i
            
            #essentially finds the smallest distance in the unsorted region, then moves it to smallest index in unsorted region, expanding the sorted region
            for j in range(i+1,len(neighbors_distances)):
                if neighbors_distances[j][1] < neighbors_distances[min_index][1]:
                    min_index = j
            
            #swap smallest dist neighbor with neighbor at i (boundry of sorted region)
            temp = neighbors_distances[i]
            neighbors_distances[i] = neighbors_distances[min_index]
            neighbors_distances[min_index] = temp

        #drop all neighbor outside top k lowest distances
        k_neighbors = neighbors_distances[:k]
        
        #majority vote
        tally_yes = 0
        tally_no = 0

        for neighbor in k_neighbors:
            label = train[neighbor[0]][0]

            if label == 1:
                tally_yes += 1
            else:
                tally_no += 1

        #in case of tie, yes wins 
        if tally_yes >= tally_no:
            pred = 1 
        else: 
            pred = 0

        real_predicted.append([int(pred), int(test_row[0])])

    print("done\n")

    return real_predicted

#Part 3 functions
def evaluate(real_predicted, title):
    """
    Evaluates KNN predictions by comparing predicted labels to real labels.
    
    Expects real_predicted as a list of [predicted, real] pairs.
    Computes and prints:
      - Confusion matrix values: TP, FP, FN, TN
      - Accuracy, Sensitivity (Recall), Specificity, Precision
    Returns a dictionary of all computed metrics.
    """

    TP = FP = FN = TN = 0

    print(f"\n{'='*55}")
    print(f"  {title}")
    print(f"{'='*55}")
    print(f"{'Real Label':<20} {'Predicted Label'}")
    print(f"{'-'*35}")

    for pair in real_predicted:
        pred = pair[0]
        real = pair[1]
        print(f"{real:<20} {pred}")

        if pred == 1 and real == 1:
            TP += 1
        elif pred == 1 and real == 0:
            FP += 1
        elif pred == 0 and real == 1:
            FN += 1
        elif pred == 0 and real == 0:
            TN += 1

    total = TP + FP + FN + TN
    accuracy    = (TP + TN) / total        if total > 0        else 0
    sensitivity = TP / (TP + FN)          if (TP + FN) > 0    else 0
    specificity = TN / (TN + FP)          if (TN + FP) > 0    else 0
    precision   = TP / (TP + FP)          if (TP + FP) > 0    else 0

    print(f"\n  Confusion Matrix:")
    print(f"    TP: {TP}  |  FP: {FP}")
    print(f"    FN: {FN}  |  TN: {TN}")
    print(f"\n  Metrics:")
    print(f"    Accuracy:           {accuracy:.4f}")
    print(f"    Sensitivity/Recall: {sensitivity:.4f}")
    print(f"    Specificity:        {specificity:.4f}")
    print(f"    Precision:          {precision:.4f}")
    print(f"{'='*55}\n")

    return {
        "title":       title,
        "TP":          TP,
        "FP":          FP,
        "FN":          FN,
        "TN":          TN,
        "accuracy":    accuracy,
        "sensitivity": sensitivity,
        "specificity": specificity,
        "precision":   precision
    }


# Part 4 Requirements
import numpy as np

#chosen framework
from sklearn.svm import SVC
from sklearn.model_selection import KFold
#feature scaling addition suggestion
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline

#Part 4 functions

#helper function to convert my returned values to be used in the evaluate() function
def combine_real_pred(y_true, y_pred):
    combined = []
    for real, pred in zip(y_true, y_pred):
        combined.append([int(pred), int(real)])
    return combined

'''
Part 4 - Support Vector Machine (SVM) Implementation Copy
- implemented a SVM with chosen framework: scikit-learn (SVC)
'''
#first need to split data into features and labels
    # label @ column 0, features @ column 1
def split_features_labels(rows):
    X=[]    # all feature columns
    y= []   # the final target label column
    
    for row in rows:
        X.append(row[1:])  # all remaining columns are our features
        y.append(int(row[0]))     #only first column is label
        
    return np.array(X), np.array(y)

# first condition - 90% training & 10% testing
# second condition - 10% training & 90% testing
#running SVM directly on the provided split training and testing datsets
def run_svm_train_test(train_rows, test_rows, title):

    #separating rows to features (X) and labels (y)
    X_train, y_train = split_features_labels(train_rows)
    X_test, y_test = split_features_labels(test_rows)

    #building the SVM model (scikit-learn)
    #model = SVC(kernel='rbf', C=1.0)
    
    #building the SVM pipeline with feature scaling
    model = Pipeline([
   ('scaler', StandardScaler()),
   ('svm', SVC(kernel='rbf', C=1.0))
])
    
    #using training data to train the model
    model.fit(X_train, y_train)

    #using testing data to predict labels
    y_pred = model.predict(X_test)
    
    #y_test = the REAL labels from the dataset
    #y_pred = the PREDICTED labels from SVM
    return y_test, y_pred
        
#printing comparison with part 3 (knn)
def compare(knn_eval, svm_eval, label):

    print("\nEvaluation Metrics:", label)
    print("Accuracy: KNN = ", knn_eval["accuracy"], "SVM = ", svm_eval["accuracy"])
    print("Sensitivity: KNN = ", knn_eval["sensitivity"], "SVM = ", svm_eval["sensitivity"])
    print("Specificity: KNN = ", knn_eval["specificity"], "SVM = ", svm_eval["specificity"])
    print("Precision: KNN = ", knn_eval["precision"], "SVM = ", svm_eval["precision"])
    
    # TP, FP, FN, and TN all in a nice formatted table for both knn and svm
    print("\nConfusion Matrix Comparative Analysis")
    print(f"{'Model':<10}{'TP':<8}{'FP':<8}{'FN':<8}{'TN':<8}")
    print("-" * 40)

    print(f"{'KNN':<10}{knn_eval['TP']:<8}{knn_eval['FP']:<8}{knn_eval['FN']:<8}{knn_eval['TN']:<8}")
    print(f"{'SVM':<10}{svm_eval['TP']:<8}{svm_eval['FP']:<8}{svm_eval['FN']:<8}{svm_eval['TN']:<8}")

if __name__ == '__main__':
    
    
    #load data (part 1)
    train,heading = load_XLSX("Training dataset.xlsx")
    test = load_XLSX("Testing dataset.xlsx")[0]

    train = process_dataset(train)
    test = process_dataset(test)
    
    #use knn on dataset (part 2)
    results_1 = perform_KNN(train,test,3)
        
    #evaluate model (Part 3)
    eval_1 = evaluate(results_1, "KNN - 90/10 Split")

    #SVM (Part 4)
    
    # first condition - 90% training & 10% testing on same data as knn! (fold = 1)
    # function call
    svm_true_90_10, svm_pred_90_10 = run_svm_train_test (
    train, test, "SVM 90% Training & 10% Testing")
    
    # second condition - 10% training & 90% testing (fold = 1)
    # function call
    svm_true_10_90, svm_pred_10_90 = run_svm_train_test (
        test, train, "SVM 10% Training & 90% Testing")
    
    #Input code here to compute accuracy, etc results of part 4 svm implementation...
    svm_eval_90_10 = evaluate(combine_real_pred(svm_true_90_10, svm_pred_90_10), "SVM - 90/10 Split")
    svm_eval_10_90 = evaluate(combine_real_pred(svm_true_10_90, svm_pred_10_90), "SVM - 10/90 Split")

    #calling function to compare results of part 3 and part 4
    compare(eval_1, svm_eval_90_10, "90/10 Split")
    compare(eval_1, svm_eval_10_90, "10/90 Split")
    
    print("Program Finished")
    
